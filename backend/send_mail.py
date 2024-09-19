"""
This module provides functions for sending emails to employees and managers.
"""

import json
import os
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Any

import openpyxl
from dotenv import load_dotenv

load_dotenv()

MAIL_SERVER: str = str(os.getenv("MAIL_SERVER"))
MAIL_PORT: int = int(str(os.getenv("MAIL_PORT")))
MAIL_USERNAME: str = str(os.getenv("MAIL_USERNAME"))
MAIL_PASSWORD: str = str(os.getenv("MAIL_PASSWORD"))
MAIL_SERVER_PROVIDER: str = str(os.getenv("MAIL_SERVER_PROVIDER"))
APP_URL: str = str(os.getenv("APP_URL"))

WORKBOOK_NAME = "MasterSheet_1.xlsx"
SHEET_NAME = "Sheet1"


def get_email_body(row: dict[str, Any]) -> str:
    body_text = """
    <html>
    <head>
    <style>

    table, th, td
    {
    border: 1px solid black;
    border-collapse: collapse:
    style="font-family calibri";
    text-align:left;
    }
    th , td {
    padding-left : 4px ;
    }

    th{
    text-align:left;
    }

    </style>

    <body>
    <table style="width:100%">
    """
    body_text += """
    <tr style="background-color:yellowgreen ; color:white">
    <th>Key</th>
    <th>Value</th>
    </tr>
    """
    for key, value in row.items():
        body_text += f"""
        <tr>
        <td>{key}</td>
        <td>{value}</td>
        </tr>
        """
    body_text += "</table>"
    button_text = "Click here to go to the app"
    button_link = f"{APP_URL}/dashboard"
    button_html = f'<a href="{button_link}" style="background-color: #4CAF50; color: white; padding: 10px 20px; text-decoration: none; border-radius: 4px;">{button_text}</a>'
    body_text += f"{button_html}"
    body_text += "</body></html>"
    return body_text


def send_mail(row: dict[str, Any]) -> bool:
    """
    Sends reminder emails to employees and weekly update emails to managers.

    Args:
        row (dict): A dictionary containing the data for a specific row.

    Returns:
        bool: True if the emails were sent successfully, False otherwise.
    """
    receivers_email = row["Mail"]
    subject = "Atlas Copco Project Status Update"
    body_text = get_email_body(row)

    # Create MIMEText object
    msg = MIMEMultipart()
    msg.attach(MIMEText(body_text, "html"))
    # Set email headers
    msg["Subject"] = subject
    msg["From"] = MAIL_USERNAME
    msg["To"] = receivers_email
    # Send the mail
    smtp_obj = smtplib.SMTP(MAIL_SERVER, 25)
    smtp_obj.connect(MAIL_SERVER, MAIL_PORT)
    # smtpObj.set_debuglevel(1)
    smtp_obj.ehlo(MAIL_SERVER_PROVIDER)
    smtp_obj.starttls()
    smtp_obj.ehlo(MAIL_SERVER_PROVIDER)
    smtp_obj.login(MAIL_USERNAME, MAIL_PASSWORD)
    # Send the email using smtplib
    send_mail_status = smtp_obj.sendmail(
        MAIL_USERNAME, receivers_email, msg.as_string()
    )
    print(f"Sending email to {receivers_email}")
    if send_mail_status:
        print(
            f"There was a problem sending email to {receivers_email}: {send_mail_status}"
        )
        return False
    return True


def send_aggregated_mail_to_all() -> None:
    """
    Sends aggregated emails to all employees and managers based on the data in the Excel file.

    Parameters:
    None

    Returns:
    None
    """
    wb = openpyxl.load_workbook(WORKBOOK_NAME)
    sheet = wb[SHEET_NAME]

    employees = {}
    manager = {}

    for r in range(2, sheet.max_row + 1):
        designation = sheet.cell(row=r, column=2).value
        task_status = sheet.cell(row=r, column=7).value
        if designation == "Employee" and (
            task_status == "DELAYED" or task_status == "INPROCESS"
        ):
            email = sheet.cell(row=r, column=9).value
            project_id = sheet.cell(row=r, column=1).value
            task_id = sheet.cell(row=r, column=4).value
            emp_id = sheet.cell(row=r, column=3).value
            task_status = sheet.cell(row=r, column=7).value

            due_date_cell = sheet.cell(row=r, column=6)
            due_date_value = due_date_cell.value
            due_date = datetime.strptime(str(due_date_value), "%Y-%m-%d %H:%M:%S")
            days_remaining = (due_date - datetime.now()).days + 1

            count_cell = sheet.cell(row=r, column=8)
            mail_count = count_cell.value
            if isinstance(mail_count, int):
                count_cell.value = mail_count + 1
            wb.save("MasterSheet_1.xlsx")

            if days_remaining < 0 and task_status == "INPROCESS":
                sheet.cell(row=r, column=7).value = "DELAYED"
                task_status = sheet.cell(row=r, column=7).value
                wb.save("MasterSheet_1.xlsx")

            if email not in employees:
                employees[email] = []

            employees[email].append(
                {
                    "emp_id": emp_id,
                    "project_id": project_id,
                    "task_id": task_id,
                    "reminders": count_cell.value,
                    "task_status": task_status,
                    "due_date": due_date,
                    "days_remaining": days_remaining,
                }
            )

        elif designation == "Manager":
            email = sheet.cell(row=r, column=9).value
            project_id = sheet.cell(row=r, column=1).value

            if email not in manager:
                manager[email] = []
            manager[email].append(project_id)

    # Print the result
    for employee_mail, employee_info_list in employees.items():
        print(f"Employee Email: {employee_mail}")
        for emp_info in employee_info_list:
            emp_id = emp_info["emp_id"]
            due_date = emp_info["due_date"]
            task_status = emp_info["task_status"]
            task_id = emp_info["task_id"]
            reminders = emp_info["reminders"]
            project_id = emp_info["project_id"]
            print(
                f"  emp_id: {emp_id}, \
                    Due Date: {due_date}, \
                    task_id: {task_id}, \
                    Reminders: {reminders}, \
                    project_id: {project_id}, \
                    Task Status: {task_status}"
            )

    # Create a dictionary to store managers and their employees
    managers_and_employees = {}

    # Populate the dictionary based on the existing data
    for employee_mail, employee_info_list in employees.items():
        for emp_info in employee_info_list:
            project_id = emp_info["project_id"]

            for email, proj_id in manager.items():
                if project_id in proj_id:
                    manager_email = email

                    if manager_email not in managers_and_employees:
                        managers_and_employees[manager_email] = []

                    managers_and_employees[manager_email].append(
                        {
                            "emp_id": emp_info["emp_id"],
                            "due_date": emp_info["due_date"],
                            "task_id": emp_info["task_id"],
                            "reminders": emp_info["reminders"],
                            "project_id": emp_info["project_id"],
                            "task_status": emp_info["task_status"],
                            "days_remaining": emp_info["days_remaining"],
                        }
                    )
                    break

    # Print the result
    # for manager_email, employees_list in managers_and_employees.items():
    #     print(f'Manager Email: {manager_email}')
    #     for emp_info in employees_list:
    #         print(f'emp_id: {emp_info["emp_id"]}, '
    #               f'Due Date: {emp_info["due_date"]}, '
    #               f'task_id: {emp_info["task_id"]}, '
    #               f'Reminders: {emp_info["reminders"]}, '
    #               f'project_id: {emp_info["project_id"]}, '
    #               f'Task Status: {emp_info["task_status"]}, '
    #               f'Days Remaining: {emp_info["days_remaining"]}')
    #     print()

    smtp_obj = smtplib.SMTP(MAIL_SERVER, MAIL_PORT)
    smtp_obj.ehlo()
    smtp_obj.starttls()
    smtp_obj.login(MAIL_USERNAME, MAIL_PASSWORD)

    # Send email to employees
    for employee_mail, employee_info_list in employees.items():
        employee_count = len(employee_info_list)
        emp_id = "abc"
        due_date = "abc"
        task_id = "abc"
        reminders = 0
        project_id = "abc"
        days_remaining = 0
        verb = "are"
        if employee_count == 1:
            verb = "is"

        subject = "Task Reminder"
        body = f"This is just a quick note to remind you that \
            {employee_count} out of your total tasks {verb} remaining.\n\n"

        for emp_info in employee_info_list:
            emp_id = emp_info["emp_id"]
            due_date = emp_info["due_date"]
            task_id = emp_info["task_id"]
            reminders = emp_info["reminders"]
            project_id = emp_info["project_id"]
            task_status = emp_info["task_status"]
            days_remaining = emp_info["days_remaining"]

            # Add information to the body
            if days_remaining < 0:
                days_remaining = 0
            body += f" Project ID: {project_id}\n \
                Task ID: {task_id}\n \
                Due Date: {due_date}\n \
                Task Status: {task_status}\n \
                Reminder No: {reminders}\n \
                Days Remaining: {days_remaining}\n\n"

        body += "\nBest Regards,\nAtlas Copco"
        # print(body)

        # Create MIMEText object
        msg = MIMEMultipart()
        msg.attach(MIMEText(body, "plain"))

        # Set email headers
        msg["From"] = "ATLAS COPCO <shreyag0104@outlook.com>"
        msg["To"] = employee_mail
        msg["Subject"] = subject

        # Send the email using smtplib
        sendmail_status = smtp_obj.sendmail(
            "shreyag0104@outlook.com", employee_mail, msg.as_string()
        )
        print(f"Sending email to {employee_mail}...")
        if sendmail_status:
            print(
                f"There was a problem sending email to {employee_mail}: {sendmail_status}"
            )

    # Send email to manager
    for manager_email, employees_list in managers_and_employees.items():
        # print(f'Manager Email: {manager_email}')
        employee_count = len(employees_list)
        emp_id = "abc"
        due_date = "abc"
        task_id = "abc"
        reminders = 0
        project_id = "abc"
        verb = "are"
        if employee_count == 1:
            verb = "is"

        subject = "Weekly Update"
        body = "This is just a quick note to update you about the remaining \
            tasks.\n\n"

        for emp_info in employees_list:
            emp_id = emp_info["emp_id"]
            due_date = emp_info["due_date"]
            task_id = emp_info["task_id"]
            reminders = emp_info["reminders"]
            project_id = emp_info["project_id"]
            task_status = emp_info["task_status"]

            body += f" Project ID: {project_id}\n \
                Task ID: {task_id}\n \
                Task Status: {task_status}\n \
                Due Date: {due_date}\n \
                Reminders Sent: {reminders}\n\n"

        body += "\nBest Regards,\nAtlas Copco"
        # print(body)

        # Create MIMEText object
        msg = MIMEMultipart()
        msg.attach(MIMEText(body, "plain"))

        # Set email headers
        msg["From"] = f"ATLAS COPCO <{MAIL_USERNAME}>"
        msg["To"] = manager_email
        msg["Subject"] = subject

        # Send the email using smtplib
        sendmail_status = smtp_obj.sendmail(
            MAIL_USERNAME, manager_email, msg.as_string()
        )
        print(f"Sending email to {manager_email}...")
        if sendmail_status:
            print(
                f"There was a problem sending email to {manager_email}: {sendmail_status}"
            )
